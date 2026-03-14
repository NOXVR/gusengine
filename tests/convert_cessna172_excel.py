"""
Convert Cessna 172 Top 100 Issues Excel → accuracy_corpus + stress_test_corpus JSON
Excel columns: #, Category, Technical Issue, Primary Complaint, Paraphrase 1-3, Validated Fix
"""
import openpyxl, json, re, os, sys

sys.stdout.reconfigure(encoding='utf-8')

EXCEL = r"I:\Backups\lilypixel\Cessna_172_Skyhawk_1969-1976_Top_100_Issues_Database.xlsx"
OUT_DIR = r"j:\GusEngine\tests"

wb = openpyxl.load_workbook(EXCEL)
ws = wb["Cessna 172 1969-76 Issues DB"]

# Category → system prefix mapping
CATEGORY_PREFIX = {}

def make_id(category, num):
    """Generate a short ID like eng_001, fuel_002, etc."""
    prefix_map = {
        "Engine": "eng",
        "Fuel": "fuel",
        "Ignition": "ign",
        "Electrical": "elec",
        "Landing Gear": "gear",
        "Flight Controls": "ctrl",
        "Exhaust": "exh",
        "Propeller": "prop",
        "Airframe": "afr",
        "Instruments": "inst",
        "Avionics": "avio",
        "Heating": "heat",
        "Cabin": "cab",
        "Hydraulic": "hyd",
        "Vacuum": "vac",
        "Pitot": "pitot",
    }
    cat_lower = category.lower()
    prefix = "misc"
    for key, val in prefix_map.items():
        if key.lower() in cat_lower:
            prefix = val
            break
    
    if prefix not in CATEGORY_PREFIX:
        CATEGORY_PREFIX[prefix] = 0
    CATEGORY_PREFIX[prefix] += 1
    return f"{prefix}_{CATEGORY_PREFIX[prefix]:03d}"

# Extract keywords from validated fix text
def extract_keywords(fix_text, category):
    """Pull important technical keywords from the validated fix."""
    keywords = set()
    # Common aviation terms to look for
    terms = [
        "compression", "cylinder", "valve", "piston", "ring", "gasket",
        "magneto", "spark plug", "ignition", "timing", "harness",
        "carburetor", "fuel", "strainer", "gascolator", "mixture",
        "alternator", "voltage", "battery", "ammeter", "circuit",
        "oil", "filter", "pump", "seal", "bearing",
        "cable", "pulley", "bellcrank", "control", "trim",
        "flap", "aileron", "elevator", "rudder",
        "nose gear", "shimmy", "dampener", "tire", "brake",
        "exhaust", "muffler", "heater", "shroud",
        "propeller", "spinner", "governor",
        "vacuum", "gyro", "pitot", "static",
        "annual", "inspection", "overhaul", "TBO",
        "Lycoming", "O-320", "Service Manual",
    ]
    fix_lower = fix_text.lower()
    for term in terms:
        if term.lower() in fix_lower:
            keywords.add(term.lower())
    return sorted(list(keywords))[:8]  # Cap at 8 keywords

accuracy_queries = []
stress_queries = []
paraphrase_groups = []

print("=== Converting Cessna 172 Excel to Corpus ===")

for row in range(2, ws.max_row + 1):
    num = ws.cell(row, 1).value
    if num is None:
        continue
    
    category = str(ws.cell(row, 2).value or "").strip()
    technical_issue = str(ws.cell(row, 3).value or "").strip()
    primary_complaint = str(ws.cell(row, 4).value or "").strip()
    paraphrase1 = str(ws.cell(row, 5).value or "").strip()
    paraphrase2 = str(ws.cell(row, 6).value or "").strip()
    paraphrase3 = str(ws.cell(row, 7).value or "").strip()
    validated_fix = str(ws.cell(row, 8).value or "").strip()
    
    if not primary_complaint or not validated_fix:
        print(f"  SKIP row {row}: missing complaint or fix")
        continue
    
    qid = make_id(category, num)
    keywords = extract_keywords(validated_fix, category)
    
    # Clean up unicode
    for char_from, char_to in [("–", "-"), ("—", "--"), ("\u2018", "'"), ("\u2019", "'"), ("\u201c", '"'), ("\u201d", '"')]:
        technical_issue = technical_issue.replace(char_from, char_to)
        primary_complaint = primary_complaint.replace(char_from, char_to)
        validated_fix = validated_fix.replace(char_from, char_to)
        paraphrase1 = paraphrase1.replace(char_from, char_to)
        paraphrase2 = paraphrase2.replace(char_from, char_to)
        paraphrase3 = paraphrase3.replace(char_from, char_to)
    
    # Accuracy corpus entry (for LLM grading)
    accuracy_entry = {
        "id": qid,
        "system": category,
        "category": "symptom_diagnosis",
        "difficulty": "medium",
        "technical_issue": technical_issue,
        "query": primary_complaint,
        "validated_fix": validated_fix,
        "known_answer_keywords": keywords,
    }
    accuracy_queries.append(accuracy_entry)
    
    # Stress test corpus entry (for stress_test_runner.py)
    stress_entry = {
        "id": qid,
        "system": category,
        "category": "symptom_diagnosis",
        "difficulty": "medium",
        "query": primary_complaint,
        "expected_systems": [category],
        "known_answer_keywords": keywords,
        "follow_ups": [],
        "technical_issue": technical_issue,
        "validated_fix": validated_fix,
    }
    stress_queries.append(stress_entry)
    
    # Paraphrase group
    variants = []
    for i, para in enumerate([paraphrase1, paraphrase2, paraphrase3], 1):
        if para and para != "None":
            variants.append({
                "id": f"{qid}_p{i}",
                "query": para,
            })
    if variants:
        paraphrase_groups.append({
            "scenario_id": qid,
            "system": category,
            "expected_keywords": keywords,
            "variants": variants,
        })
    
    print(f"  [{num:3d}] {qid}: {technical_issue[:60]}...")

# Write accuracy corpus
accuracy_corpus = {
    "metadata": {
        "vehicle_id": "1975_cessna_172_skyhawk_aviation",
        "version": "1.0.0",
        "description": "Cessna 172 Skyhawk (1969-1976) Top 100 Issues - A&P/IA verified",
        "total_queries": len(accuracy_queries),
        "source": "Cessna_172_Skyhawk_1969-1976_Top_100_Issues_Database.xlsx",
    },
    "queries": accuracy_queries,
}

accuracy_path = os.path.join(OUT_DIR, "accuracy_corpus_cessna172.json")
with open(accuracy_path, "w", encoding="utf-8") as f:
    json.dump(accuracy_corpus, f, indent=4, ensure_ascii=False)
print(f"\nAccuracy corpus: {accuracy_path} ({len(accuracy_queries)} queries)")

# Write stress test corpus
stress_corpus = {
    "metadata": {
        "vehicle_id": "1975_cessna_172_skyhawk_aviation",
        "version": "1.0.0",
        "description": "Cessna 172 Skyhawk (1969-1976) stress test corpus with paraphrases",
        "total_queries": len(stress_queries),
    },
    "queries": stress_queries,
    "paraphrase_groups": paraphrase_groups,
}

stress_path = os.path.join(OUT_DIR, "stress_test_corpus_cessna172.json")
with open(stress_path, "w", encoding="utf-8") as f:
    json.dump(stress_corpus, f, indent=4, ensure_ascii=False)
print(f"Stress test corpus: {stress_path} ({len(stress_queries)} queries + {len(paraphrase_groups)} paraphrase groups)")

# Summary by category
print("\n=== Category Distribution ===")
cats = {}
for q in accuracy_queries:
    cat = q["system"]
    cats[cat] = cats.get(cat, 0) + 1
for cat, count in sorted(cats.items(), key=lambda x: -x[1]):
    print(f"  {cat}: {count}")

"""
Convert 'MerCruiser 5.0/5.7 Top 100 Issues' Excel → stress_test_corpus_mercruiser.json

Produces the exact JSON format expected by stress_test_runner.py v3.
Keywords are auto-extracted from the 'Validated Fix' column using a
MerCruiser GM V8 sterndrive-specific component vocabulary.
"""
import json
import os
import re
import sys

try:
    import openpyxl
except ImportError:
    print("ERROR: openpyxl required. Install with: pip install openpyxl")
    sys.exit(1)

# ── MerCruiser GM V8 Sterndrive Component Vocabulary ──────────────────
MERCRUISER_COMPONENTS = [
    # Engine — GM 305/350 Small Block
    "intake manifold", "exhaust manifold", "head gasket", "cylinder head",
    "crankshaft", "camshaft", "piston", "piston ring", "connecting rod",
    "valve cover", "valve", "valve stem seal", "rocker arm", "lifter",
    "oil pump", "oil pan", "oil cooler", "oil pressure", "oil filter",
    "timing chain", "timing cover", "harmonic balancer",
    "rear main seal", "front seal", "freeze plug", "core plug",
    "engine mount", "motor mount", "flywheel", "flexplate",
    "compression", "bore", "stroke",
    "PCV", "breather",

    # Cooling — Dual Circuit (Raw + Freshwater)
    "raw water pump", "seawater pump", "impeller", "water pump",
    "thermostat", "heat exchanger", "radiator",
    "circulating pump", "freshwater pump",
    "coolant hose", "expansion tank", "coolant recovery",
    "temperature sender", "temperature sensor", "overheat alarm",
    "water intake", "seacock", "raw water filter", "strainer",
    "anode", "zinc", "pencil zinc",

    # Exhaust — Manifolds & Risers
    "exhaust riser", "exhaust elbow", "riser gasket",
    "exhaust manifold", "water jacket", "water passage",
    "exhaust bellows", "exhaust hose",
    "catalytic converter",

    # Fuel System
    "carburetor", "flame arrestor", "fuel pump", "fuel filter",
    "fuel line", "fuel tank", "fuel sender", "fuel gauge",
    "fuel injector", "throttle body", "TBI",
    "choke", "idle mixture", "accelerator pump",
    "anti-siphon valve", "fuel water separator",
    "fuel pressure regulator",

    # Ignition / Electrical
    "distributor", "distributor cap", "rotor", "ignition coil",
    "spark plug", "spark plug wire", "ignition module",
    "Thunderbolt", "EST", "electronic ignition",
    "alternator", "voltage regulator", "starter", "starter solenoid",
    "battery", "battery switch", "battery cable",
    "MerCathode", "galvanic isolator",
    "fuse", "relay", "wiring harness", "ground strap",
    "shift interrupt switch", "neutral safety switch",
    "trim sender", "oil pressure sender",
    "overheat horn", "alarm buzzer",

    # Sterndrive Unit
    "sterndrive", "outdrive", "Alpha One", "Alpha", "Bravo",
    "upper gearcase", "lower gearcase", "gearcase",
    "drive shaft", "prop shaft", "propeller shaft",
    "forward gear", "reverse gear", "clutch dog",
    "shift shaft", "shift cable", "shift spool",
    "U-joint", "universal joint", "gimbal bearing",
    "bearing carrier", "prop shaft seal",
    "water pump impeller", "drive impeller",
    "gear oil", "gear lube", "drain plug",
    "propeller", "prop nut", "tab washer",
    "trim tab", "sacrificial anode",

    # Transom Assembly
    "transom plate", "transom assembly",
    "exhaust bellows", "U-joint bellows", "shift cable bellows",
    "bellows", "bellows clamp", "hose clamp",
    "gimbal housing", "gimbal ring",
    "engine coupler", "coupling",
    "water hose", "water pickup",

    # Trim / Tilt
    "trim pump", "tilt pump", "trim motor",
    "trim cylinder", "tilt cylinder", "trim ram",
    "trim sender", "trim gauge",
    "trim limit switch", "tilt switch",
    "hydraulic fluid", "hydraulic oil",
    "trim hose", "trim fitting",

    # Corrosion / Zincs / MerCathode
    "zinc anode", "sacrificial anode", "pencil zinc",
    "MerCathode", "galvanic corrosion", "electrolysis",
    "stray current", "bonding wire",

    # Winterization
    "antifreeze", "fogging oil", "fuel stabilizer",
    "drain plug", "block drain", "manifold drain",
    "water separator",
]

# ── Category → System mapping ──────────────────────────────────────────
CATEGORY_TO_SYSTEM = {
    "Exhaust Manifolds & Risers": "Exhaust",
    "Cooling System": "Cooling",
    "Engine (GM 305/350)": "Engine",
    "Fuel System": "Fuel",
    "Ignition / Electrical": "Electrical",
    "Sterndrive Unit": "Sterndrive",
    "Transom Assembly": "Transom",
    "Trim / Tilt System": "Trim",
    "Corrosion / Zincs": "Corrosion",
    "Winterization Failures": "Winterization",
}

CATEGORY_TO_PREFIX = {
    "Exhaust Manifolds & Risers": "exh",
    "Cooling System": "cool",
    "Engine (GM 305/350)": "eng",
    "Fuel System": "fuel",
    "Ignition / Electrical": "elec",
    "Sterndrive Unit": "drive",
    "Transom Assembly": "trans",
    "Trim / Tilt System": "trim",
    "Corrosion / Zincs": "corr",
    "Winterization Failures": "winter",
}


def extract_keywords(validated_fix: str) -> list[str]:
    """Extract MerCruiser-specific component keywords from validated fix text."""
    if not validated_fix:
        return []
    fix_lower = validated_fix.lower()
    found = []
    for component in MERCRUISER_COMPONENTS:
        if component.lower() in fix_lower:
            found.append(component)
    seen = set()
    unique = []
    for kw in found:
        if kw not in seen:
            seen.add(kw)
            unique.append(kw)
    return unique


def main():
    xlsx_path = r"I:\Backups\lilypixel\MerCruiser_5.0_5.7_Top_Issues_Database.xlsx"
    if not os.path.exists(xlsx_path):
        print(f"ERROR: Excel file not found: {xlsx_path}")
        sys.exit(1)

    wb = openpyxl.load_workbook(xlsx_path)
    ws = wb.active

    queries = []
    paraphrase_groups = []
    cat_counters = {}

    for row_num in range(2, ws.max_row + 1):
        cells = [ws.cell(row=row_num, column=c).value for c in range(1, 9)]
        num, category, technical_issue, primary_complaint, para1, para2, para3, validated_fix = cells

        if not category or not primary_complaint:
            continue

        prefix = CATEGORY_TO_PREFIX.get(category, "misc")
        cat_counters[prefix] = cat_counters.get(prefix, 0) + 1
        query_id = f"mc_{prefix}_{cat_counters[prefix]:03d}"

        system = CATEGORY_TO_SYSTEM.get(category, category)
        keywords = extract_keywords(validated_fix or "")

        query_entry = {
            "id": query_id,
            "system": system,
            "category": "symptom_diagnosis",
            "difficulty": "medium",
            "query": primary_complaint.strip(),
            "expected_systems": [system],
            "known_answer_keywords": keywords[:8],
            "technical_issue": (technical_issue or "").strip(),
            "validated_fix": (validated_fix or "").strip(),
            "follow_ups": [],
        }
        queries.append(query_entry)

        variants = []
        for i, para_text in enumerate([para1, para2, para3], 1):
            if para_text and para_text.strip():
                variants.append({
                    "id": f"{query_id}_p{i}",
                    "query": para_text.strip(),
                })

        if variants:
            paraphrase_groups.append({
                "scenario_id": query_id,
                "system": system,
                "expected_keywords": keywords[:6],
                "variants": variants,
            })

    corpus = {
        "metadata": {
            "vehicle_id": "2000_mercury_mercruiser_marine",
            "version": "1.0.0",
            "description": "Full decision-tree accuracy corpus for MerCruiser 5.0L/5.7L GM V-8 sterndrive marine engine",
            "total_queries": len(queries),
            "total_paraphrases": sum(len(pg["variants"]) for pg in paraphrase_groups),
            "source": "MerCruiser_5.0_5.7_Top_Issues_Database.xlsx",
        },
        "queries": queries,
        "paraphrase_groups": paraphrase_groups,
    }

    script_dir = os.path.dirname(os.path.abspath(__file__))
    output_path = os.path.join(script_dir, "stress_test_corpus_mercruiser.json")
    with open(output_path, "w", encoding="utf-8") as f:
        json.dump(corpus, f, indent=4, ensure_ascii=False)

    print(f"Converted {len(queries)} queries + {len(paraphrase_groups)} paraphrase groups")
    print(f"Output: {output_path}")

    systems = {}
    total_kw = 0
    for q in queries:
        s = q["system"]
        systems[s] = systems.get(s, 0) + 1
        total_kw += len(q["known_answer_keywords"])

    print(f"\nQueries by system:")
    for s, count in sorted(systems.items(), key=lambda x: -x[1]):
        print(f"  {s}: {count}")
    print(f"\nAvg keywords per query: {total_kw / len(queries):.1f}")


if __name__ == "__main__":
    main()

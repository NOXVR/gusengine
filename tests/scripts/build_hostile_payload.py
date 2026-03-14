"""
build_hostile_payload.py — Generate hostile analysis payloads from GusEngine test data.

Reads the cleaned XLSX (answer key), test run JSONL (Gus's output), and scorecard
(grading) for a given vehicle corpus, then produces batched text files ready to paste
into adversarial LLMs.

Usage:
    python build_hostile_payload.py --vehicle mustang [--partial-only] [--batch-size 10]
    python build_hostile_payload.py --vehicle mercedes [--partial-only]
    python build_hostile_payload.py --vehicle mercruiser [--partial-only]
    python build_hostile_payload.py --vehicle cessna [--partial-only]
"""

import argparse
import json
import re
import openpyxl
from pathlib import Path

# Vehicle configurations
VEHICLES = {
    'mustang': {
        'xlsx': r'I:\Backups\lilypixel\1965_Ford_Mustang_Issues_Cleaned.xlsx',
        'jsonl': r'J:\GusEngine\tests\results\run_20260309_163545.jsonl',
        'scorecard': r'J:\GusEngine\tests\results\accuracy_scorecard.md',
        'name': '1965 Ford Mustang',
    },
    'mercedes': {
        'xlsx': r'I:\Backups\lilypixel\1976_Mercedes_450SL_Issues_Cleaned.xlsx',
        'jsonl': r'J:\GusEngine\tests\results\run_20260311_190826.jsonl',
        'scorecard': r'J:\GusEngine\tests\results\accuracy_scorecard_450sl.md',
        'name': '1976 Mercedes-Benz 450SL',
    },
    'mercruiser': {
        'xlsx': r'I:\Backups\lilypixel\MerCruiser_5.0_5.7_Issues_Cleaned.xlsx',
        'jsonl': r'J:\GusEngine\tests\results\run_20260311_225329.jsonl',
        'scorecard': r'J:\GusEngine\tests\results\accuracy_scorecard_mercruiser.md',
        'name': 'MerCruiser 5.0L/5.7L GM V-8',
    },
    'cessna': {
        'xlsx': r'I:\Backups\lilypixel\Cessna_172_Skyhawk_1969-1976_Top_100_Issues_Database.xlsx',
        'jsonl': r'J:\GusEngine\tests\results\run_20260312_133607.jsonl',
        'scorecard': r'J:\GusEngine\tests\results\accuracy_scorecard_cessna172.md',
        'name': '1975 Cessna 172 Skyhawk',
    },
}

OUTPUT_DIR = Path(r'J:\GusEngine\tests\results\hostile_payloads')


def load_xlsx(path):
    """Load the cleaned XLSX answer key. Returns dict of {issue_num: {category, issue, fix}}."""
    wb = openpyxl.load_workbook(path)
    ws = wb.active
    issues = {}
    for row in range(2, ws.max_row + 1):
        num = ws.cell(row, 1).value
        if num is None:
            continue
        try:
            num = int(num)
        except (ValueError, TypeError):
            continue
        issues[num] = {
            'category': ws.cell(row, 2).value or '',
            'issue': ws.cell(row, 3).value or '',
            'complaint': ws.cell(row, 4).value or '',
            'fix': ws.cell(row, 8).value or '',
        }
    return issues


def load_jsonl(path):
    """Load the test run JSONL. Returns dict of {query_id: gus_output_text}.
    
    The JSONL contains multiple entries per query_id (BFS Decision Tree branches).
    We take the first entry (initial triage) and extract the diagnostic content
    from response_parsed fields.
    """
    results = {}
    with open(path, 'r', encoding='utf-8') as f:
        for line in f:
            line = line.strip()
            if not line:
                continue
            try:
                record = json.loads(line)
            except json.JSONDecodeError:
                continue
            
            qid = record.get('query_id', '')
            
            # Only take the first entry per query_id (initial triage)
            if qid in results:
                continue
            
            # Extract diagnostic content from response_parsed
            parsed = record.get('response_parsed', {})
            if not parsed or not isinstance(parsed, dict):
                # Fallback: try response_raw as a string
                raw = record.get('response_raw', '')
                if raw and isinstance(raw, str):
                    try:
                        parsed = json.loads(raw)
                    except (json.JSONDecodeError, TypeError):
                        parsed = {}
            
            parts = []
            
            # Current state / phase
            state = parsed.get('current_state', '')
            if state:
                parts.append(f"[Phase: {state}]")
            
            # Diagnostic reasoning
            reasoning = parsed.get('diagnostic_reasoning', '')
            if reasoning:
                parts.append(f"DIAGNOSTIC REASONING:\n{reasoning}")
            
            # Mechanic instructions
            instructions = parsed.get('mechanic_instructions', '')
            if instructions:
                parts.append(f"MECHANIC INSTRUCTIONS:\n{instructions}")
            
            # Answer path prompts (the diagnostic options presented to user)
            answer_paths = parsed.get('answer_path_prompts', [])
            if answer_paths and isinstance(answer_paths, list):
                paths_text = "DIAGNOSTIC OPTIONS PRESENTED:"
                for i, ap in enumerate(answer_paths, 1):
                    if isinstance(ap, dict):
                        label = ap.get('label', ap.get('text', str(ap)))
                        paths_text += f"\n  Option {i}: {label}"
                    elif isinstance(ap, str):
                        paths_text += f"\n  Option {i}: {ap}"
                parts.append(paths_text)
            
            gus_output = '\n\n'.join(parts) if parts else '[NO OUTPUT CAPTURED]'
            
            # Truncate very long outputs to keep payload manageable
            if len(gus_output) > 3000:
                gus_output = gus_output[:3000] + '\n[...TRUNCATED FOR BREVITY...]'
            
            results[qid] = gus_output
    return results



def load_scorecard(path):
    """Load the grading scorecard. Returns dict of {issue_num: {query_id, issue, grade, rationale}}."""
    grades = {}
    with open(path, 'r', encoding='utf-8') as f:
        content = f.read()
    
    # Parse the results table
    # Format: | # | Query ID | Technical Issue | Grade | Rationale |
    pattern = r'\|\s*(\d+)\s*\|\s*(\S+)\s*\|\s*(.+?)\s*\|\s*(CORRECT|PARTIALLY_CORRECT|INCORRECT|NOT_APPLICABLE)\s*\|\s*(.+?)\s*\|'
    for match in re.finditer(pattern, content):
        num = int(match.group(1))
        grades[num] = {
            'query_id': match.group(2).strip(),
            'issue': match.group(3).strip(),
            'grade': match.group(4).strip(),
            'rationale': match.group(5).strip(),
        }
    return grades


def load_prompt_template():
    """Load the hostile analysis prompt template."""
    template_path = Path(r'J:\GusEngine\tests\scripts\hostile_analysis_prompt.md')
    with open(template_path, 'r', encoding='utf-8') as f:
        return f.read()


def format_issue_payload(num, xlsx_data, gus_output, grade_data):
    """Format a single issue into the hostile analysis payload."""
    return f"""### ISSUE #{num}: {xlsx_data['issue']}
**Category:** {xlsx_data['category']}
**Customer Complaint:** {xlsx_data['complaint']}

**ANSWER KEY FIX (Ground Truth):**
{xlsx_data['fix']}

**GUS'S DIAGNOSTIC OUTPUT:**
{gus_output}

**ORIGINAL GRADE:** {grade_data['grade']}
**GRADING RATIONALE:** {grade_data['rationale']}

---
"""


def build_payloads(vehicle_key, partial_only=False, batch_size=10):
    """Build hostile analysis payloads for a vehicle."""
    config = VEHICLES[vehicle_key]
    
    print(f"Loading data for {config['name']}...")
    xlsx_data = load_xlsx(config['xlsx'])
    jsonl_data = load_jsonl(config['jsonl'])
    scorecard = load_scorecard(config['scorecard'])
    template = load_prompt_template()
    
    print(f"  XLSX: {len(xlsx_data)} issues")
    print(f"  JSONL: {len(jsonl_data)} queries")
    print(f"  Scorecard: {len(scorecard)} grades")
    
    # Filter issues
    issues_to_audit = []
    for num, grade_data in sorted(scorecard.items()):
        if grade_data['grade'] == 'NOT_APPLICABLE':
            continue
        if partial_only and grade_data['grade'] != 'PARTIALLY_CORRECT':
            continue
        
        # Match JSONL by query_id
        query_id = grade_data['query_id']
        gus_output = jsonl_data.get(query_id, '[JSONL ENTRY NOT FOUND]')
        
        if num not in xlsx_data:
            print(f"  WARNING: Issue #{num} not in XLSX, skipping")
            continue
        
        issues_to_audit.append((num, xlsx_data[num], gus_output, grade_data))
    
    mode_label = "PARTIAL_ONLY" if partial_only else "FULL"
    print(f"\n  Issues to audit ({mode_label}): {len(issues_to_audit)}")
    
    if not issues_to_audit:
        print("  No issues to audit!")
        return
    
    # Create output directory
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    
    # Batch the issues
    batches = []
    for i in range(0, len(issues_to_audit), batch_size):
        batch = issues_to_audit[i:i + batch_size]
        batches.append(batch)
    
    print(f"  Generating {len(batches)} batch(es) of up to {batch_size} issues each...\n")
    
    for batch_idx, batch in enumerate(batches, 1):
        # Build the issues payload
        issues_text = ""
        for num, xlsx, gus, grade in batch:
            issues_text += format_issue_payload(num, xlsx, gus, grade)
        
        # Insert into template
        payload = template.replace('{ISSUES_PAYLOAD}', issues_text)
        
        # Add vehicle context at the top
        header = f"**Vehicle:** {config['name']}\n**Batch:** {batch_idx}/{len(batches)} ({mode_label})\n**Issues in this batch:** {', '.join(f'#{n}' for n, _, _, _ in batch)}\n\n"
        payload = header + payload
        
        # Write to file
        suffix = 'partial' if partial_only else 'full'
        filename = f"hostile_{vehicle_key}_{suffix}_batch{batch_idx:02d}.txt"
        filepath = OUTPUT_DIR / filename
        with open(filepath, 'w', encoding='utf-8') as f:
            f.write(payload)
        
        issue_nums = [n for n, _, _, _ in batch]
        grades = [g['grade'] for _, _, _, g in batch]
        print(f"  Batch {batch_idx}: Issues {issue_nums}")
        print(f"    Grades: {grades}")
        print(f"    -> {filepath}")
        print(f"    Size: {len(payload):,} chars")
        print()


def main():
    parser = argparse.ArgumentParser(description='Build hostile analysis payloads')
    parser.add_argument('--vehicle', required=True, choices=list(VEHICLES.keys()),
                        help='Vehicle corpus to audit')
    parser.add_argument('--partial-only', action='store_true',
                        help='Only audit PARTIALLY_CORRECT grades (the weakest points)')
    parser.add_argument('--batch-size', type=int, default=10,
                        help='Issues per batch (default: 10)')
    args = parser.parse_args()
    
    build_payloads(args.vehicle, args.partial_only, args.batch_size)
    print("Done!")


if __name__ == '__main__':
    main()

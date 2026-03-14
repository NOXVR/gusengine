"""
Convert '1976 Mercedes 450SL Top 100 Issues' Excel → stress_test_corpus_450sl.json

Produces the exact JSON format expected by stress_test_runner.py v3:
  - queries[]: id, system, category, difficulty, query, expected_systems,
               known_answer_keywords, technical_issue, validated_fix, follow_ups
  - paraphrase_groups[]: scenario_id, system, expected_keywords, variants[]

Keywords are auto-extracted from the 'Validated Fix' column using a
Mercedes M117/R107-specific component vocabulary.
"""
import json
import os
import re
import sys

try:
    import openpyxl
except ImportError:
    print("ERROR: openpyxl required. Install with: pip install openpyxl")
    sys.exit(1)

# ── Mercedes M117 / R107 Component Vocabulary ──────────────────────────
# Used to auto-extract ground-truth keywords from the Validated Fix text.
MERCEDES_COMPONENTS = [
    # Engine — M117 V8
    "timing chain", "chain guide", "chain tensioner", "timing cover",
    "camshaft", "crankshaft", "connecting rod", "piston", "piston ring",
    "cylinder head", "head gasket", "valve", "valve stem seal", "valve cover gasket",
    "valve spring", "rocker arm", "hydraulic lifter", "cam lobe",
    "oil pump", "oil pan gasket", "oil filter", "oil cooler", "oil pressure",
    "rear main seal", "front crank seal", "pilot bearing",
    "engine mount", "motor mount", "flexplate", "flywheel",
    "PCV", "breather", "dipstick tube",
    "compression", "bore", "stroke",

    # Fuel — Bosch K-Jetronic CIS
    "K-Jetronic", "CIS", "fuel distributor", "fuel injector",
    "warm control pressure regulator", "control pressure",
    "system pressure", "rest pressure",
    "cold start valve", "thermo-time switch",
    "auxiliary air valve", "auxiliary air slide",
    "fuel accumulator", "fuel pump", "fuel filter", "fuel pump relay",
    "mixture screw", "CO adjustment", "air flow sensor", "sensor plate",
    "fuel pressure", "fuel line", "fuel tank",

    # Ignition
    "distributor", "distributor cap", "rotor", "ignition coil",
    "ignition points", "condenser", "spark plug",
    "ballast resistor", "ignition module", "ignition switch",
    "timing", "advance", "vacuum advance", "centrifugal advance",
    "firing order", "dwell",

    # Cooling
    "thermostat", "water pump", "radiator", "fan clutch",
    "heater core", "heater valve", "coolant hose", "freeze plug",
    "expansion tank", "coolant temperature sensor",
    "auxiliary fan", "fan shroud",

    # Electrical
    "alternator", "voltage regulator", "starter", "starter solenoid",
    "battery", "fuse box", "relay", "headlight relay",
    "turn signal switch", "combination switch", "wiper motor",
    "instrument cluster", "gauge", "sending unit",
    "ground strap", "wiring harness",

    # Brakes
    "brake booster", "master cylinder", "wheel cylinder",
    "brake caliper", "brake disc", "brake rotor", "brake pad",
    "brake drum", "brake shoe", "brake line", "brake hose",
    "parking brake", "parking brake cable",
    "ABS", "proportioning valve",

    # Transmission — 722 Automatic
    "automatic transmission", "torque converter",
    "vacuum modulator", "kickdown", "downshift",
    "shift valve", "governor", "valve body",
    "transmission mount", "transmission fluid",
    "front pump seal", "rear seal", "pan gasket",
    "flex disc", "driveshaft", "center bearing",

    # Suspension / Steering
    "steering box", "drag link", "idler arm", "tie rod",
    "ball joint", "control arm", "subframe",
    "shock absorber", "spring", "sway bar", "stabilizer bar",
    "wheel bearing", "hub",

    # Rear Axle
    "differential", "axle shaft", "CV joint",
    "rear axle seal", "pinion seal",

    # Exhaust / Emissions
    "exhaust manifold", "catalytic converter", "muffler",
    "EGR", "air pump", "check valve",
    "lambda", "oxygen sensor",

    # Climate / HVAC
    "AC compressor", "evaporator", "condenser",
    "blower motor", "heater valve",
    "vacuum actuator", "climate control",

    # Body / Convertible / Interior
    "door hinge", "window regulator", "door lock",
    "weatherstrip", "windshield seal", "trunk seal",
    "convertible top", "hydraulic cylinder",
    "floor pan", "rocker panel", "fender",
    "headliner", "carpet", "seat",
    "dashboard", "speedometer cable",
]

# ── Category → System mapping ──────────────────────────────────────────
CATEGORY_TO_SYSTEM = {
    "Engine (M117 V8)": "Engine",
    "Fuel System (Bosch CIS)": "Fuel",
    "Ignition System": "Ignition",
    "Cooling System": "Cooling",
    "Electrical System": "Electrical",
    "Brakes": "Brakes",
    "Transmission (722 Auto)": "Transmission",
    "Suspension / Steering": "Suspension",
    "Driveshaft / Rear Axle": "Drivetrain",
    "Exhaust / Emissions": "Exhaust",
    "Climate Control / HVAC": "Climate",
    "Body / Rust": "Body",
    "Interior": "Interior",
    "Convertible Top / Hardtop": "Convertible",
    "Weatherstripping / Seals": "Seals",
}


def extract_keywords(validated_fix: str) -> list[str]:
    """Extract Mercedes-specific component keywords from validated fix text."""
    if not validated_fix:
        return []
    fix_lower = validated_fix.lower()
    found = []
    for component in MERCEDES_COMPONENTS:
        if component.lower() in fix_lower:
            found.append(component)
    # Deduplicate while preserving order
    seen = set()
    unique = []
    for kw in found:
        if kw not in seen:
            seen.add(kw)
            unique.append(kw)
    return unique


def category_to_id_prefix(category: str) -> str:
    """Convert category name to a short ID prefix."""
    prefixes = {
        "Engine (M117 V8)": "eng",
        "Fuel System (Bosch CIS)": "fuel",
        "Ignition System": "ign",
        "Cooling System": "cool",
        "Electrical System": "elec",
        "Brakes": "brk",
        "Transmission (722 Auto)": "trans",
        "Suspension / Steering": "susp",
        "Driveshaft / Rear Axle": "drive",
        "Exhaust / Emissions": "exh",
        "Climate Control / HVAC": "hvac",
        "Body / Rust": "body",
        "Interior": "int",
        "Convertible Top / Hardtop": "conv",
        "Weatherstripping / Seals": "seal",
    }
    return prefixes.get(category, "misc")


def main():
    xlsx_path = r"I:\Backups\lilypixel\1976_Mercedes_450SL_Top_100_Issues_Database.xlsx"
    if not os.path.exists(xlsx_path):
        print(f"ERROR: Excel file not found: {xlsx_path}")
        sys.exit(1)

    wb = openpyxl.load_workbook(xlsx_path)
    ws = wb.active

    queries = []
    paraphrase_groups = []

    # Track per-category counters for IDs
    cat_counters = {}

    for row_num in range(2, ws.max_row + 1):
        cells = [ws.cell(row=row_num, column=c).value for c in range(1, 9)]
        num, category, technical_issue, primary_complaint, para1, para2, para3, validated_fix = cells

        if not category or not primary_complaint:
            continue

        # Generate ID
        prefix = category_to_id_prefix(category)
        cat_counters[prefix] = cat_counters.get(prefix, 0) + 1
        query_id = f"m450_{prefix}_{cat_counters[prefix]:03d}"

        system = CATEGORY_TO_SYSTEM.get(category, category)
        keywords = extract_keywords(validated_fix or "")

        # Main query entry
        query_entry = {
            "id": query_id,
            "system": system,
            "category": "symptom_diagnosis",
            "difficulty": "medium",
            "query": primary_complaint.strip(),
            "expected_systems": [system],
            "known_answer_keywords": keywords[:8],  # Cap at 8 keywords
            "technical_issue": (technical_issue or "").strip(),
            "validated_fix": (validated_fix or "").strip(),
            "follow_ups": [],
        }
        queries.append(query_entry)

        # Paraphrase group (if paraphrases exist)
        variants = []
        for i, para_text in enumerate([para1, para2, para3], 1):
            if para_text and para_text.strip():
                variants.append({
                    "id": f"{query_id}_p{i}",
                    "query": para_text.strip(),
                })

        if variants:
            paraphrase_groups.append({
                "scenario_id": query_id,
                "system": system,
                "expected_keywords": keywords[:6],
                "variants": variants,
            })

    # Build corpus
    corpus = {
        "metadata": {
            "vehicle_id": "1976_mercedes_450sl",
            "version": "1.0.0",
            "description": "Full decision-tree accuracy corpus for 1976 Mercedes-Benz 450SL (M117 V8, Bosch K-Jetronic CIS, 722 auto)",
            "total_queries": len(queries),
            "total_paraphrases": sum(len(pg["variants"]) for pg in paraphrase_groups),
            "source": "1976_Mercedes_450SL_Top_100_Issues_Database.xlsx",
        },
        "queries": queries,
        "paraphrase_groups": paraphrase_groups,
    }

    # Write output
    script_dir = os.path.dirname(os.path.abspath(__file__))
    output_path = os.path.join(script_dir, "stress_test_corpus_450sl.json")
    with open(output_path, "w", encoding="utf-8") as f:
        json.dump(corpus, f, indent=4, ensure_ascii=False)

    # Summary
    print(f"Converted {len(queries)} queries + {len(paraphrase_groups)} paraphrase groups")
    print(f"Output: {output_path}")

    # Stats
    systems = {}
    total_kw = 0
    for q in queries:
        s = q["system"]
        systems[s] = systems.get(s, 0) + 1
        total_kw += len(q["known_answer_keywords"])

    print(f"\nQueries by system:")
    for s, count in sorted(systems.items(), key=lambda x: -x[1]):
        print(f"  {s}: {count}")
    print(f"\nAvg keywords per query: {total_kw / len(queries):.1f}")


if __name__ == "__main__":
    main()
